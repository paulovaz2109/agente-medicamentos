# -*- coding: utf-8 -*-
import os
import urllib.request
import urllib.parse
from datetime import date, datetime
from openpyxl import load_workbook

CAMINHO_PLANILHA = "estoque.xlsx"
DIAS_ALERTA      = 5
ABA_ESTOQUE      = "Estoque"
COLUNA_PACIENTE  = 2
COLUNA_MED       = 3
COLUNA_DURACAO   = 5

MEU_NUMERO   = os.getenv("MEU_NUMERO")
APIKEY       = os.getenv("CALLMEBOT_APIKEY")
MEU_NUMERO2  = os.getenv("MEU_NUMERO2", "")
APIKEY2      = os.getenv("CALLMEBOT_APIKEY2", "")


def ler_planilha():
    wb = load_workbook(CAMINHO_PLANILHA, data_only=True)
    ws = wb[ABA_ESTOQUE]
    hoje = date.today()
    alertas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        paciente = row[COLUNA_PACIENTE - 1]
        med      = row[COLUNA_MED - 1]
        duracao  = row[COLUNA_DURACAO - 1]
        if not med or not duracao:
            continue
        if isinstance(duracao, str):
            continue
        if isinstance(duracao, datetime):
            duracao = duracao.date()
        if not isinstance(duracao, date):
            continue
        dias_restantes = (duracao - hoje).days
        if dias_restantes <= DIAS_ALERTA:
            alertas.append({
                "paciente":       str(paciente) if paciente else "?",
                "medicamento":    str(med),
                "duracao":        duracao.strftime("%d/%m/%Y"),
                "dias_restantes": dias_restantes,
            })
    alertas.sort(key=lambda x: x["dias_restantes"])
    return alertas


def formatar_mensagem(alertas):
    hoje_str = date.today().strftime("%d/%m/%Y")
    linhas = [
        "*ALERTA DE ESTOQUE - " + hoje_str + "*",
        "Medicamentos que vencem em ate " + str(DIAS_ALERTA) + " dias:\n",
    ]
    for a in alertas:
        d = a["dias_restantes"]
        if d < 0:    status = "VENCIDO"
        elif d == 0: status = "ACABA HOJE"
        elif d == 1: status = "1 dia restante"
        else:        status = str(d) + " dias restantes"
        linhas.append(
            "* *" + a["paciente"] + " - " + a["medicamento"] + "*\n"
            "  Dura ate: *" + a["duracao"] + "* (" + status + ")"
        )
    linhas.append("\n_Agente automatico - Estoque Medicamentos_")
    return "\n".join(linhas)


def enviar_callmebot(numero, apikey, mensagem):
    texto = urllib.parse.quote(mensagem)
    url = ("https://api.callmebot.com/whatsapp.php"
           "?phone=" + numero +
           "&text=" + texto +
           "&apikey=" + apikey)
    try:
        req = urllib.request.urlopen(url, timeout=15)
        print("OK - Enviado para " + numero)
        return True
    except Exception as e:
        print("ERRO ao enviar para " + numero + ": " + str(e))
        return False


def main():
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")
    print("[" + agora + "] Iniciando verificacao de estoque...")

    alertas = ler_planilha()

    if not alertas:
        mensagem = (
            "*ESTOQUE OK - " + date.today().strftime("%d/%m/%Y") + "*\n"
            "Nenhum medicamento acaba nos proximos " + str(DIAS_ALERTA) + " dias.\n\n"
            "_Agente automatico - Estoque Medicamentos_"
        )
        print("Nenhum alerta critico. Enviando mensagem de status...")
    else:
        print("ATENCAO - " + str(len(alertas)) + " medicamento(s) critico(s):")
        for a in alertas:
            print("   -> " + a["paciente"] + " | " + a["medicamento"] + " | " + str(a["dias_restantes"]) + " dias")
        mensagem = formatar_mensagem(alertas)

    print("\nMensagem:\n" + mensagem)

    enviar_callmebot(MEU_NUMERO, APIKEY, mensagem)
    if MEU_NUMERO2 and APIKEY2:
        enviar_callmebot(MEU_NUMERO2, APIKEY2, mensagem)

    print("Concluido.")


if __name__ == "__main__":
    main()
